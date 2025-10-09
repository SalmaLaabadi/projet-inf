import io
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Les entêtes fixes du fichier Excel
DEFAULT_HEADERS = [
    "CIN", "Nom", "Prenom", "Departement",
    "SessionID", "Intitule", "Date",
    "Present", "NoteSur20", "Verdict"  # Apte/Inapte
]

def create_header_only_workbook(date_obj, headers=None, sheet_name="Feuille1"):
    """
    Crée un fichier Excel (.xlsx) contenant uniquement les entêtes de colonnes.
    Exemple de colonnes : CIN, Nom, Prenom, etc.
    """
    headers = headers or DEFAULT_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Ligne d'entête
    ws.append(headers)

    # Style (gras + centré)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Largeur auto
    for col_index, title in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_index).column_letter
        ws.column_dimensions[col_letter].width = max(12, len(title) + 2)

    # Sauvegarde en mémoire
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    # Nom du fichier
    filename = f"tests_{date_obj.isoformat()}_entete.xlsx"
    return ContentFile(mem.read(), name=filename)
