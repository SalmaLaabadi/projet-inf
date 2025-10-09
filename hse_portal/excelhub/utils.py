import io
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Entêtes par défaut
DEFAULT_HEADERS = [
    "CIN", "Nom", "Prenom", "Departement",
    "SessionID", "Intitule", "Date",
    "Present", "NoteSur20", "Verdict"  # Apte/Inapte
]

def create_header_only_workbook(date_obj, headers=None, sheet_name="Feuille1"):
    headers = headers or DEFAULT_HEADERS
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ligne d'entête
    ws.append(headers)

    # style entête
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # largeur colonnes
    for i, title in enumerate(headers, start=1):
        col = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col].width = max(12, len(title) + 2)

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    filename = f"tests_{date_obj.isoformat()}_entete.xlsx"
    return ContentFile(mem.read(), name=filename)
